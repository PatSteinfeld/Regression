import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile

@st.cache_data(show_spinner=False)
def unmerge_excel(file):
    """
    Opens the Excel file (from a BytesIO stream), unmerges all cells by copying the value
    from the top-left cell of each merged range into all cells in that range,
    and then saves the cleaned workbook to a temporary file.
    Returns the temporary file path.
    """
    # Load workbook from file-like object (BytesIO)
    wb = load_workbook(file, data_only=True)
    for sheet in wb.worksheets:
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = merged_range.bounds
            # Get value from top-left cell
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            # Fill each cell in the merged range with this value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value
            # Unmerge the cells
            sheet.unmerge_cells(str(merged_range))
    # Save the cleaned workbook to a temporary file
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name

@st.cache_data(show_spinner=False)
def extract_data_from_cleaned_excel(file):
    """
    Preprocess the Excel file by unmerging cells, then extract required data from the PAP sheet.
    
    Extraction Logic (based on manual inspection):
    - Audit Types: Row 7 (index 6), starting at column index 1 onward.
    - Number of Man-Days: Row 8 (index 7), starting at column index 1; only numeric values are kept.
    - Proposed Audit Dates: Row 9 (index 8), starting at column index 1.
    - Sites: Assumed to be in column index 10 from row 12 onward.
    - Activities: Assumed to be in column index 12 adjacent to the site rows.
    """
    # Preprocess the file (unmerge cells)
    cleaned_file = unmerge_excel(file)
    
    # Read the PAP sheet without header inference (to use custom indexing)
    pap_df = pd.read_excel(cleaned_file, sheet_name="PAP", header=None)
    
    # --- Extract Audit Details ---
    # Audit Types: row index 6, starting from column index 1
    audit_types_raw = pap_df.iloc[6, 1:].dropna().tolist()
    audit_types = [x for x in audit_types_raw if isinstance(x, str) and "insert" not in x]
    
    # Man-Days: row index 7, starting from column index 1; keep only numeric values
    man_days_raw = pap_df.iloc[7, 1:].dropna().tolist()
    man_days = [x for x in man_days_raw if isinstance(x, (int, float))]
    
    # Proposed Audit Dates: row index 8, starting from column index 1; filter out extraneous text
    audit_dates_raw = pap_df.iloc[8, 1:].dropna().tolist()
    audit_dates = [x for x in audit_dates_raw if isinstance(x, str) and "Site" not in x]
    
    # --- Extract Sites and Activities ---
    # Assume Sites are in column index 10, from row index 12 to 30
    sites_raw = pap_df.iloc[12:30, 10].dropna().tolist()
    sites = sites_raw
    
    # Activities: Assume they are in column index 12 (adjacent to the site rows)
    activities_dict = {}
    for idx, site in enumerate(sites_raw):
        try:
            activity = pap_df.iloc[12 + idx, 12]
            if pd.notna(activity):
                activities_dict[site] = activity
            else:
                activities_dict[site] = "No activity specified"
        except Exception:
            activities_dict[site] = "No activity specified"
    
    return {
        "audit_types": audit_types,
        "man_days": man_days,
        "audit_dates": audit_dates,
        "sites": sites,
        "activities": activities_dict
    }

def main():
    st.title("Audit Planner (PAP Sheet) - Improved Extraction")
    st.write("This app preprocesses a complex Excel file (with merged cells and multiple headers) to extract audit details.")
    
    # 1️⃣ File Upload
    uploaded_file = st.file_uploader("Upload the Excel File", type=["xlsm", "xlsx"])
    if uploaded_file:
        data = extract_data_from_cleaned_excel(uploaded_file)
        
        st.subheader("Extracted Data")
        st.write("**Audit Types:**", data["audit_types"])
        st.write("**Number of Man-Days:**", data["man_days"])
        st.write("**Proposed Audit Dates:**", data["audit_dates"])
        st.write("**Sites:**", data["sites"])
        
        # 2️⃣ User Inputs
        num_auditors = st.selectbox("Select Number of Auditors", list(range(1, 11)))
        
        auditor_names_input = st.text_input("Enter Auditor Names (comma separated)")
        auditor_names = [name.strip() for name in auditor_names_input.split(",") if name.strip()]
        
        selected_audit_type = st.selectbox("Select Audit Type", data["audit_types"])
        selected_site = st.selectbox("Select Site", data["sites"])
        selected_date = st.date_input("Select Audit Date", datetime.today())
        
        # Display activity for the selected site (if available)
        activity_for_site = data["activities"].get(selected_site, "No activity specified")
        st.write("**Activity for selected site:**", activity_for_site)
        
        # 3️⃣ Generate Audit Plan
        if st.button("Generate Audit Plan"):
            try:
                index = data["audit_types"].index(selected_audit_type)
                total_man_days = data["man_days"][index] if index < len(data["man_days"]) else None
            except Exception:
                total_man_days = None
            
            total_hours = total_man_days * 8 if total_man_days is not None else 0
            
            # Create a simple schedule DataFrame
            plan_data = {
                "Audit Type": [selected_audit_type],
                "Site": [selected_site],
                "Activity": [activity_for_site],
                "Time Allocation (hours)": [total_hours],
                "Audit Date": [selected_date],
                "Assigned Auditor": [auditor_names[0] if auditor_names else "Not Assigned"]
            }
            plan_df = pd.DataFrame(plan_data)
            
            st.subheader("Audit Planning Schedule")
            st.dataframe(plan_df)
            st.success("Audit plan generated successfully!")
    else:
        st.info("Please upload an Excel file to get started.")

if __name__ == "__main__":
    main()




















